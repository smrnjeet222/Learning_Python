import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
       wb = xl.load_workbook(filename)
       sheet = wb['Sheet1']

       sheet['d1'].value = 'Correct Prices'

       for row in range(2, sheet.max_row + 1):
              cell = sheet.cell(row, 3)
              correction_price = cell.value * 0.9
              correction_price_cell = sheet.cell(row , 4)
              correction_price_cell.value = f'${correction_price}'

#       values = Reference(sheet, min_row= 2, max_row= 4, min_col= 4, max_col= 4)

#       chart =  BarChart()
#       chart.add_data(values)
#       sheet.add_chart(chart, (5,2))

       wb.save('newfile.xlsx')
